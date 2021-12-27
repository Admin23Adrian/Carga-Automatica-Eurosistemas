from sap import meteteensap
import os, os.path
import win32com.client
from os import walk
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import os
from getpass import getuser
import shutil
import openpyxl
from datetime import date
from datetime import datetime
import win32com.client as win32
from os import remove
import pythoncom
from datetime import datetime
import requests
import sys
import subprocess
import time
from openpyxl import load_workbook


def eurosistemas():
    
    user = getuser()
    root = Tk()
    dt = datetime.now()    # Fecha y hora actual
    day = dt.day
    month = dt.month
    year = dt.year
    hoy =  str(day) + "-" + str(month) + "-" + str(year)
    root.title("Pedidos_Eurosistemas")
    root.resizable(0,0)
    root.geometry('300x300+500+50'.format(500, 600))

    
    miFrame=Frame(root,width=500)
    miFrame.pack()
    
    miFrame2=Frame(root)
    miFrame2.pack()

    def lector_excel():
        """
        Funcion que se encarga de leer todos los datos del excel y devolverlos en forma de listas
        para posteriormente usarlas.
        """
        user=getuser()
        dt = datetime.now()
        day = dt.day
        month = dt.month
        year = dt.year
        hoy=str(day) + "-" + str(month) + "-" + str(year)

        pythoncom.CoInitialize()

        path = ('C:/Users/' + user +  '/Desktop/EUROSISTEMAS' + '/' + "eurosistemas.xlsx")
        path2 = ('C:/Users/' + user +  '/Desktop/EUROSISTEMAS' + '/' + "Euro1.xlsx")
    
        shutil.copy(path, path2)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["inicio"]

        ultimafiladelws = len(ws['I']) + 1

        # LISTAS POR COMPRENSION. CODIGO MAS LIMPIO Y LEGIBLE.
        # obs_internas = [ws[f"A{i}"].value for i in range(9, ultimafiladelws) if ws[f"A{i}"].value != None]
        codificacion_afiliado = [ws[f"B{i}"].value for i in range(9, ultimafiladelws) if ws[f"B{i}"].value != None]
        cantidades = [ws[f"D{i}"].value for i in range(9, ultimafiladelws) if ws[f"D{i}"].value != None]
        # nro_pedidos_ext = [ws[f"E{i}"].value for i in range(9, ultimafiladelws) if ws[f"E{i}"].value != None]
        nom_ape_afiliados = [ws[f"G{i}"].value for i in range(9, ultimafiladelws) if ws[f"G{i}"].value != None]
        farmacias = [ws[f"I{i}"].value for i in range(9, ultimafiladelws) if ws[f"I{i}"].value != None]
        direcciones_farmacia = [ws[f"J{i}"].value for i in range(9, ultimafiladelws) if ws[f"J{i}"].value != None]
        localidades_farmacia = [ws[f"K{i}"].value for i in range(9, ultimafiladelws) if ws[f"K{i}"].value != None]
        dispones = [ws[f"L{i}"].value for i in range(9, ultimafiladelws) if ws[f"L{i}"].value != None]
        idafiliados_sap = [ws[f"M{i}"].value for i in range(9, ultimafiladelws) if ws[f"M{i}"].value != None]
        materiales_sap = [ws[f"N{i}"].value for i in range(9, ultimafiladelws) if ws[f"N{i}"].value != None]
        convenios = [ws[f"O{i}"].value for i in range(9, ultimafiladelws) if ws[f"O{i}"].value != None]
        clases_pedidos = [ws[f"Q{i}"].value for i in range(9, ultimafiladelws) if ws[f"Q{i}"].value != None]
        almacenes = [ws[f"R{i}"].value for i in range(9, ultimafiladelws) if ws[f"R{i}"].value != None]
        solicitantes = [ws[f"S{i}"].value for i in range(9, ultimafiladelws) if ws[f"S{i}"].value != None]
        fechas_entregas = [ws[f"T{i}"].value for i in range(9, ultimafiladelws) if ws[f"T{i}"].value != None]
        centros = [ws[f"U{i}"].value for i in range(9, ultimafiladelws) if ws[f"U{i}"].value != None]
        turnos = [ws[f"V{i}"].value for i in range(9, ultimafiladelws) if ws[f"V{i}"].value != None]
        canales = [ws[f"W{i}"].value for i in range(9, ultimafiladelws) if ws[f"W{i}"].value != None]
        sectores = [ws[f"X{i}"].value for i in range(9, ultimafiladelws) if ws[f"X{i}"].value != None]
        dni_afil = [ws[f"Y{i}"].value for i in range(9, ultimafiladelws) if ws[f"Y{i}"].value != None]
        sexo = [ws[f"Z{i}"].value for i in range(9, ultimafiladelws)]
        wb.close()
        # return cantidades, nom_ape_afiliados, nro_pedidos_ext, farmacias, direcciones_farmacia, localidades_farmacia, dispones, idafiliados_sap, materiales_sap, convenios, clases_pedidos, materiales_sap, convenios, clases_pedidos, almacenes, solicitantes, fechas_entregas, centros, turnos, canales, sectores, codificacion_afiliado, dni_afil, sexo
        return cantidades, nom_ape_afiliados, farmacias, direcciones_farmacia, localidades_farmacia, dispones, idafiliados_sap, materiales_sap, convenios, clases_pedidos, materiales_sap, convenios, clases_pedidos, almacenes, solicitantes, fechas_entregas, centros, turnos, canales, sectores, codificacion_afiliado, dni_afil, sexo

    def generador_pedidos():
        
        user = getuser()
        dt = datetime.now()
        day = dt.day
        month = dt.month
        year = dt.year
        hoy = str(day) + "-" + str(month) + "-" + str(year)
        
        # LISTAS TRAIDAS DEL EXCEL - REPRESENTAN CADA COLUMNA.
        cantidades, nom_ape_afiliados, farmacias, direcciones_farmacia, localidades_farmacia, dispones, idafiliados_sap, materiales_sap, convenios, clases_pedidos, materiales_sap, convenios, clases_pedidos, almacenes, solicitantes, fechas_entregas, centros, turnos, canales, sectores, codigos_afiliados, dni_afiliados, sexos = lector_excel()

        pythoncom.CoInitialize()
        path_copia = f"C:/Users/{user}/Desktop/EUROSISTEMAS/Euro1.xlsx"
        wb_copia = openpyxl.load_workbook(path_copia, data_only = True)
        ws_copia = wb_copia["inicio"]

        # LISTAS A CARGAR PARA CADA AFILIADO
        l_cantidades = []
        l_materiales_sap = []
        # -------------------------------- #

        codigo_afiliado_anterior = 0
        material_anterior = 0
        filas = []
        
        for i in range(0, len(idafiliados_sap) + 1):
            """
            Este for se para en una fila y las condicionales evaluan si la fila actual donde esta parado es
            igual a la anterior. Si la condicion se cumple significa que tenemos el mismo afiliado con
            mas de una posicion de medicacion. Si no se cumple la condicion, el afiliado es diferente y 
            se factura siempre el afiliado anterior.
            """
            if i != len(idafiliados_sap):
                codigo_afiliado = codigos_afiliados[i]
                medicacion = materiales_sap[i]
                cantidad_medicacion = cantidades[i]
            else:
                print(f"FACTURANDO AF: {codigo_afiliado_anterior}")
                print(f"Cantidades: {l_cantidades} | Materiales: {l_materiales_sap}")
                nombre_y_apellido = nom_ape_afiliados[i - 1]
                nombre = nombre_y_apellido.split()[0]
                apellido = nombre_y_apellido.split()[1]
                datos_farmacia = farmacias[i - 1] + " | " + direcciones_farmacia[i - 1] + " | " + localidades_farmacia[i - 1]
                # IR A SAP A CREAR EL PEDIDO EN CASO QUE SEA LA ULTIMA FILA
                # ultimo_pedido = meteteensap(0, clases_pedidos[i - 1], canales[i - 1], sectores[i - 1], nro_pedidos_ext[i - 1], solicitantes[i - 1], dispones[i - 1], fechas_entregas[i - 1], l_materiales_sap, centros[i - 1], l_cantidades, convenios[i - 1], turnos[i - 1], almacenes[i - 1], idafiliados_sap[i - 1], codigos_afiliados[i - 1], nombre, apellido, dni_afiliados[i - 1], datos_farmacia, sexos[i - 1])
                ultimo_pedido = meteteensap(0, clases_pedidos[i - 1], canales[i - 1], sectores[i - 1], solicitantes[i - 1], dispones[i - 1], fechas_entregas[i - 1], l_materiales_sap, centros[i - 1], l_cantidades, convenios[i - 1], turnos[i - 1], almacenes[i - 1], idafiliados_sap[i - 1], codigos_afiliados[i - 1], nombre, apellido, dni_afiliados[i - 1], datos_farmacia, sexos[i - 1])
                for fila in filas:
                    try:
                        ws_copia[f"AA{fila}"] = ultimo_pedido
                    except:
                        ws_copia[f"AA{fila}"] = f"Error."
                print(f"Se creo y se guardo el pedido: {ultimo_pedido}")
                break

            # SE COMIENZA A EVALUEAR LOS AFILIADOS. Si es el primero entra y si hay mas de una posicion para el mismo afiliado.
            if codigo_afiliado == codigo_afiliado_anterior or codigo_afiliado_anterior == 0:
                print(f"Afiliado: {codigo_afiliado}. Medicacion ID: {medicacion}. Cantidad: {cantidad_medicacion}")
                l_cantidades.append(cantidad_medicacion)
                l_materiales_sap.append(medicacion)
                filas.append(str(i + 9))
                
                codigo_afiliado_anterior = codigo_afiliado
                material_anterior = medicacion

            ## CONDICIONAL PARA EVALUAR AFILIADOS ##
            ## --> Si el afiliado donde actualmente esta parado es distinto al anterior se factura el anterior.
            elif codigo_afiliado != codigo_afiliado_anterior:
                print(f"FACTURANDO AF: {codigo_afiliado_anterior}")
                print(f"Cantidades: {l_cantidades} | Materiales: {l_materiales_sap}")
                nombre_y_apellido = nom_ape_afiliados[i - 1]
                nombre = nombre_y_apellido.split()[0]
                apellido = nombre_y_apellido.split()[1]
                # OBSERVACIOES INTERNAS. PARA CASOS DE CARGA CON DISPONE GENERICO.
                datos_farmacia = farmacias[i - 1] + " | " + direcciones_farmacia[i - 1] + " | " + localidades_farmacia[i - 1]
                # IR A SAP A CREAR EL PEDIDO.
                # pedido = meteteensap(0, clases_pedidos[i - 1], canales[i - 1], sectores[i - 1], nro_pedidos_ext[i - 1], solicitantes[i - 1], dispones[i - 1], fechas_entregas[i - 1], l_materiales_sap, centros[i - 1], l_cantidades, convenios[i - 1], turnos[i - 1], almacenes[i - 1], idafiliados_sap[i - 1], codigo_afiliado_anterior, nombre, apellido, dni_afiliados[i - 1], datos_farmacia, sexos[i - 1])
                pedido = meteteensap(0, clases_pedidos[i - 1], canales[i - 1], sectores[i - 1], solicitantes[i - 1], dispones[i - 1], fechas_entregas[i - 1], l_materiales_sap, centros[i - 1], l_cantidades, convenios[i - 1], turnos[i - 1], almacenes[i - 1], idafiliados_sap[i - 1], codigo_afiliado_anterior, nombre, apellido, dni_afiliados[i - 1], datos_farmacia, sexos[i - 1])
                
                for fila in filas:
                    try:
                        ws_copia[f"AA{fila}"] = pedido
                    except:
                        ws_copia[f"AA{fila}"] = "Error."
                print(f"Datos de carga de pedido: {pedido}")
                print()
                l_cantidades.clear()
                l_materiales_sap.clear()
                filas.clear()

                l_cantidades.append(cantidad_medicacion)
                l_materiales_sap.append(medicacion)
                filas.append(str(i + 9))

                codigo_afiliado_anterior = codigo_afiliado
                material_anterior = medicacion

        ## SE GUARDA Y SE CIERRA EL DOCUMENTO DE EXCEL. 
        wb_copia.save(path_copia)
        wb_copia.close()
    
    botonCrear = Button(miFrame, text="Ejecutar", command = generador_pedidos)
    botonCrear.grid(row = 5, column = 2, sticky = "e", padx = 10, pady = 10)

    root.mainloop()


#---------------------------------------------------------FIN-----------------------------------------------------------#
if __name__=="__main__":
    eurosistemas()






