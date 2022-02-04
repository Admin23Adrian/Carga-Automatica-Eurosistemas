import os, os.path
from os import walk
import os
from getpass import getuser
from datetime import date
from datetime import datetime
import win32com.client as win32
import pythoncom
import win32com.client
import sys
import subprocess
import time
import openpyxl
from openpyxl import load_workbook


def meteteensap(sesionsap, nro_afiliado):
    
    pythoncom.CoInitialize()

    SapGuiAuto = win32com.client.GetObject('SAPGUI')
    if not type(SapGuiAuto) == win32com.client.CDispatch:
        return

    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32com.client.CDispatch:
        SapGuiAuto = None
        return
    connection = application.Children(0)

    if not type(connection) == win32com.client.CDispatch:
        application = None
        SapGuiAuto = None
        return

    session = connection.Children(sesionsap)
    if not type(session) == win32com.client.CDispatch:
        connection = None
        application = None
        SapGuiAuto = None
        return
    
    try:
        print(f"** Consultando afiliado en SAP: {nro_afiliado}")
        session.findById("wnd[0]").maximize
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nz_sd_estados_hist"
        session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]/usr/ctxtSO_VKORG-LOW").text = "SC10"
        session.findById("wnd[0]/usr/ctxtSO_VDATU-LOW").text = ""
        session.findById("wnd[0]/usr/ctxtSO_AFILI-LOW").text = f"{nro_afiliado}"
        session.findById("wnd[0]/usr/ctxtSO_AFILI-LOW").setFocus()
        session.findById("wnd[0]/usr/ctxtSO_AFILI-LOW").caretPosition = 8
        session.findById("wnd[0]/tbar[1]/btn[8]").press()
        session.findById("wnd[0]/tbar[1]/btn[33]").press()
        session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").setCurrentCell(32,"TEXT")
        session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").firstVisibleRow = 28
        session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").selectedRows = "32"
        session.findById("wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell").clickCurrentCell()
        
        # CAMPOS QUE NECESITAMOS EXTRAER.
        cont = 0
        try:
            for i in range(0, 4000):
                dispone = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(i, "DESTINA")
                nombre_farmacia = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(i, "DESTINA_NAME")
                domicilio = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(i, "DOMICILIO")
                localidad = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(i, "LOCALIDAD")
                cont += 1
        except:
            print(f"Entrando en except de sap. {str(cont - 1)}")
            codificacion_afiliado = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "AFILIADO_NOM")
            nro_cliente = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "CLIENTE")
            nombre_cliente = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "CLIENTE_NOM")
            centro = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "VSTEL")
            turno = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "ZZTURNO")
            dispone = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "DESTINA")
            nombre_farmacia = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "DESTINA_NAME")
            domicilio = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "DOMICILIO")
            localidad = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "LOCALIDAD")
            provincia = session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell").getcellvalue(cont - 1, "PROVINCIA")
            
        return codificacion_afiliado, nro_afiliado, nro_cliente, nombre_cliente, centro, turno, dispone, nombre_farmacia, domicilio, localidad, provincia
    
    except Exception as e:
        print("Error en el script.. Programa finalizado!", e)


def completar_datos_excel(ruta_excel):

    excel = load_workbook(ruta_excel, data_only=True)
    hoja = excel["Padron Nico"]
    hoja_padron = excel["AFILIADOS_SACADOS_SAP_SIN_FARMA"]
    max_rows_excel = hoja.max_row
    try:
        print("INGRESANDO A DATOS DEL EXCEL. HOJA: Padron Nico")
        for i in range(2, max_rows_excel + 1):
            print(f"{hoja[f'D{str(i)}'].value}")
            nro_afiliado = hoja[f"A{i}"].value
            try:
                codificacion_afiliado, nro_afiliado, nro_cliente, nombre_cliente, centro, turno, dispone, nombre_farmacia, domicilio, localidad, provincia = meteteensap(0, f"{nro_afiliado}")
                print(codificacion_afiliado, nro_afiliado, nro_cliente, nombre_cliente, centro, turno, dispone, nombre_farmacia, domicilio, localidad, provincia)
                hoja_padron[f"A{i}"].value = codificacion_afiliado
                hoja_padron[f"B{i}"].value = codificacion_afiliado
                hoja_padron[f"C{i}"].value = nro_afiliado
                hoja_padron[f"D{i}"].value = nro_cliente
                hoja_padron[f"E{i}"].value = nombre_cliente
                hoja_padron[f"F{i}"].value = centro
                hoja_padron[f"G{i}"].value = turno
                hoja_padron[f"H{i}"].value = dispone
                hoja_padron[f"I{i}"].value = domicilio
                hoja_padron[f"J{i}"].value = localidad
                hoja_padron[f"K{i}"].value = provincia
                print("----------------------------------------------------------------\n")
            except:
                continue
    except Exception as e:
        print("Algo salio mal en el excel.", e)
    finally:
        excel.save(ruta_excel)
        excel.close()

ruta_excel = "C:/Users/aalarcon/Desktop/OyP/CARGA AUTOMATICA EUROSISTEMAS/P2.xlsx"
completar_datos_excel(ruta_excel)    
